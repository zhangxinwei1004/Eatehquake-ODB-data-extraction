from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple, Union

from openpyxl import Workbook, load_workbook


M_capacity_kN_m = 0.0
INPUT_EXCEL = "postprocess_results.xlsx"
TIME_HISTORY_SHEET = "Time_History"
OUTPUT_DIR = "Origin_export"

SOURCE_COLUMNS = [
    "时间_s",
    "塔顶位移_m",
    "底部位移_m",
    "塔顶相对位移_m",
    "塔顶加速度_m_s2",
    "底部加速度_m_s2",
    "底部剪力_RF1_N",
    "底部剪力_RF1_kN",
    "底部弯矩_RM1_N_m",
    "底部倾覆弯矩_RM2_N_m",
    "底部弯矩_RM3_N_m",
    "底部倾覆弯矩_RM2_kN_m",
    "底部倾覆弯矩_M_ot_N_m",
    "底部倾覆弯矩_M_ot_kN_m",
    "合成倾覆弯矩_N_m",
    "合成倾覆弯矩_kN_m",
]

Number = Union[int, float]
CellValue = Union[str, Number, None]


def to_float(value: CellValue, column_name: str, row_index: int) -> float:
    if value is None or value == "":
        raise ValueError("第 %d 行列 '%s' 为空，无法转换为数字。" % (row_index, column_name))
    number = float(value)
    if math.isnan(number) or math.isinf(number):
        raise ValueError("第 %d 行列 '%s' 的值不是有限数字：%r" % (row_index, column_name, value))
    return number


def validate_inputs(columns: Sequence[str], col_to_index: Dict[str, int]) -> None:
    missing = []
    for name in ["时间_s", "塔顶相对位移_m", "塔顶加速度_m_s2", "底部加速度_m_s2"]:
        if name not in col_to_index:
            missing.append(name)
    if "底部剪力_RF1_kN" not in col_to_index and "底部剪力_RF1_N" not in col_to_index:
        missing.append("底部剪力_RF1_kN 或 底部剪力_RF1_N")
    if not any(name in col_to_index for name in ["底部倾覆弯矩_M_ot_kN_m", "底部倾覆弯矩_RM2_kN_m", "底部倾覆弯矩_M_ot_N_m", "底部倾覆弯矩_RM2_N_m"]):
        missing.append("可用的底部倾覆弯矩列")
    if "合成倾覆弯矩_kN_m" not in col_to_index and "合成倾覆弯矩_N_m" not in col_to_index:
        missing.append("合成倾覆弯矩_kN_m 或 合成倾覆弯矩_N_m")
    if missing:
        print("当前 Excel 中实际列名：")
        for name in columns:
            print("  %s" % name)
        print("缺少以下列：")
        for name in missing:
            print("  %s" % name)
        raise RuntimeError("Excel 列名不完整，请检查表头。")


def normalize_record(record: Dict[str, float]) -> Dict[str, float]:
    if "底部剪力_RF1_kN" not in record and "底部剪力_RF1_N" in record:
        record["底部剪力_RF1_kN"] = record["底部剪力_RF1_N"] / 1000.0
    if "底部倾覆弯矩_M_ot_kN_m" not in record:
        if "底部倾覆弯矩_RM2_kN_m" in record:
            record["底部倾覆弯矩_M_ot_kN_m"] = record["底部倾覆弯矩_RM2_kN_m"]
        elif "底部倾覆弯矩_M_ot_N_m" in record:
            record["底部倾覆弯矩_M_ot_kN_m"] = record["底部倾覆弯矩_M_ot_N_m"] / 1000.0
        elif "底部倾覆弯矩_RM2_N_m" in record:
            record["底部倾覆弯矩_M_ot_kN_m"] = record["底部倾覆弯矩_RM2_N_m"] / 1000.0
    if "合成倾覆弯矩_kN_m" not in record and "合成倾覆弯矩_N_m" in record:
        record["合成倾覆弯矩_kN_m"] = record["合成倾覆弯矩_N_m"] / 1000.0
    return record


def read_time_history(excel_path: Path) -> Tuple[List[str], List[Dict[str, float]]]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if TIME_HISTORY_SHEET not in wb.sheetnames:
            print("未找到工作表：%s" % TIME_HISTORY_SHEET)
            print("当前 Excel 中所有工作表名称：")
            for name in wb.sheetnames:
                print("  %s" % name)
            raise RuntimeError("请确认工作表名称。")
        ws = wb[TIME_HISTORY_SHEET]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows)
        columns = [str(value).strip() if value is not None else "" for value in header_row]
        col_to_index = {name: idx for idx, name in enumerate(columns) if name}
        validate_inputs(columns, col_to_index)
        data = []
        for row_index, row in enumerate(rows, start=2):
            if row is None or all(value is None for value in row):
                continue
            record = {}
            for col in SOURCE_COLUMNS:
                if col in col_to_index:
                    idx = col_to_index[col]
                    record[col] = to_float(row[idx] if idx < len(row) else None, col, row_index)
            data.append(normalize_record(record))
        return columns, data
    finally:
        wb.close()


def abs_peak(data: Sequence[Dict[str, float]], column: str) -> Tuple[float, float]:
    row = max(data, key=lambda item: abs(item[column]))
    return row[column], row["时间_s"]


def write_csv_utf8_sig(path: Path, headers: Sequence[str], rows: Iterable[Sequence[CellValue]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)


def make_tables(data: Sequence[Dict[str, float]]) -> Dict[str, Tuple[List[str], List[List[CellValue]]]]:
    fig5_rows = [[row["时间_s"], row["底部倾覆弯矩_M_ot_kN_m"]] for row in data]
    m_ot_max = max(abs(row[1]) for row in fig5_rows)
    return {
        "Fig1_输入加速度": (["时间_s", "底部加速度_m_s2"], [[row["时间_s"], row["底部加速度_m_s2"]] for row in data]),
        "Fig2_塔顶相对位移": (["时间_s", "塔顶相对位移_m"], [[row["时间_s"], row["塔顶相对位移_m"]] for row in data]),
        "Fig3_加速度对比": (["时间_s", "塔顶加速度_m_s2", "底部加速度_m_s2"], [[row["时间_s"], row["塔顶加速度_m_s2"], row["底部加速度_m_s2"]] for row in data]),
        "Fig4_底部剪力": (["时间_s", "底部剪力_RF1_kN"], [[row["时间_s"], row["底部剪力_RF1_kN"]] for row in data]),
        "Fig5_倾覆弯矩": (["时间_s", "底部倾覆弯矩_M_ot_kN_m"], fig5_rows),
        "Fig6_抗倾覆对比": (["项目", "弯矩_kN_m"], [["最大倾覆弯矩", m_ot_max], ["抗倾覆能力", M_capacity_kN_m]]),
    }


def make_summary(data: Sequence[Dict[str, float]]) -> Tuple[List[str], List[List[CellValue]], Dict[str, CellValue]]:
    input_acc, input_time = abs_peak(data, "底部加速度_m_s2")
    u_rel, u_time = abs_peak(data, "塔顶相对位移_m")
    top_acc, top_time = abs_peak(data, "塔顶加速度_m_s2")
    shear, shear_time = abs_peak(data, "底部剪力_RF1_kN")
    mot, mot_time = abs_peak(data, "底部倾覆弯矩_M_ot_kN_m")
    mres, mres_time = abs_peak(data, "合成倾覆弯矩_kN_m")
    eta = abs(top_acc) / abs(input_acc) if abs(input_acc) > 0 else "底部加速度最大绝对值为0，无法计算"
    k_value = "未设置抗倾覆能力" if M_capacity_kN_m == 0.0 else M_capacity_kN_m / abs(mot)
    headers = ["统计项目", "数值", "时间_s", "单位"]
    rows = [
        ["最大输入加速度", input_acc, input_time, "m/s^2"],
        ["最大塔顶相对位移", u_rel, u_time, "m"],
        ["最大塔顶加速度", top_acc, top_time, "m/s^2"],
        ["最大底部剪力", shear, shear_time, "kN"],
        ["最大底部倾覆弯矩", mot, mot_time, "kN*m"],
        ["最大合成倾覆弯矩", mres, mres_time, "kN*m"],
        ["塔顶加速度放大系数_eta_a", eta, "", "-"],
        ["抗倾覆安全系数_K", k_value, "", "-"],
    ]
    values = {
        "最大输入加速度": input_acc,
        "最大输入加速度时间": input_time,
        "最大塔顶相对位移": u_rel,
        "最大塔顶相对位移时间": u_time,
        "最大塔顶加速度": top_acc,
        "最大塔顶加速度时间": top_time,
        "最大底部剪力": shear,
        "最大底部剪力时间": shear_time,
        "最大底部倾覆弯矩": mot,
        "最大底部倾覆弯矩时间": mot_time,
        "最大合成倾覆弯矩": mres,
        "最大合成倾覆弯矩时间": mres_time,
        "塔顶加速度放大系数": eta,
        "抗倾覆安全系数K": k_value,
    }
    return headers, rows, values


def write_workbook(path: Path, tables, summary_headers, summary_rows) -> None:
    wb = Workbook()
    first = True
    for sheet_name, (headers, rows) in tables.items():
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        ws.append(headers)
        for row in rows:
            ws.append(row)
    ws = wb.create_sheet("Summary")
    ws.append(list(summary_headers))
    for row in summary_rows:
        ws.append(list(row))
    wb.save(path)


def main() -> None:
    root = Path(__file__).resolve().parent
    excel_path = root / INPUT_EXCEL
    output_dir = root / OUTPUT_DIR
    output_dir.mkdir(exist_ok=True)
    columns, data = read_time_history(excel_path)
    tables = make_tables(data)
    summary_headers, summary_rows, values = make_summary(data)
    names = [
        ("Fig1_ElCentro_input_acceleration.csv", "Fig1_输入加速度"),
        ("Fig2_top_relative_displacement.csv", "Fig2_塔顶相对位移"),
        ("Fig3_top_base_acceleration_comparison.csv", "Fig3_加速度对比"),
        ("Fig4_base_shear_time_history.csv", "Fig4_底部剪力"),
        ("Fig5_base_overturning_moment.csv", "Fig5_倾覆弯矩"),
        ("Fig6_overturning_moment_vs_capacity.csv", "Fig6_抗倾覆对比"),
    ]
    csv_paths = []
    for filename, sheet_name in names:
        headers, rows = tables[sheet_name]
        path = output_dir / filename
        write_csv_utf8_sig(path, headers, rows)
        csv_paths.append(path)
    summary_csv = output_dir / "Summary.csv"
    write_csv_utf8_sig(summary_csv, summary_headers, summary_rows)
    workbook_path = output_dir / "Origin_plot_data.xlsx"
    write_workbook(workbook_path, tables, summary_headers, summary_rows)

    print("成功读取的 Excel 文件路径：%s" % excel_path)
    print("成功读取的工作表名称：%s" % TIME_HISTORY_SHEET)
    print("实际识别到的列名：%s" % ", ".join(columns))
    for path in csv_paths + [summary_csv]:
        print("CSV 文件输出路径：%s" % path)
    print("Origin_plot_data.xlsx 的输出路径：%s" % workbook_path)
    print("最大输入加速度及其时间：%s, %s s" % (values["最大输入加速度"], values["最大输入加速度时间"]))
    print("最大塔顶相对位移及其时间：%s, %s s" % (values["最大塔顶相对位移"], values["最大塔顶相对位移时间"]))
    print("最大塔顶加速度及其时间：%s, %s s" % (values["最大塔顶加速度"], values["最大塔顶加速度时间"]))
    print("最大底部剪力及其时间：%s, %s s" % (values["最大底部剪力"], values["最大底部剪力时间"]))
    print("最大底部倾覆弯矩及其时间：%s, %s s" % (values["最大底部倾覆弯矩"], values["最大底部倾覆弯矩时间"]))
    print("最大合成倾覆弯矩及其时间：%s, %s s" % (values["最大合成倾覆弯矩"], values["最大合成倾覆弯矩时间"]))
    print("塔顶加速度放大系数：%s" % values["塔顶加速度放大系数"])
    print("抗倾覆安全系数 K：%s" % values["抗倾覆安全系数K"])


if __name__ == "__main__":
    main()
